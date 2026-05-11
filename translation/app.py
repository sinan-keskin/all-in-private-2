import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from deep_translator import GoogleTranslator, MyMemoryTranslator, exceptions
import ssl, re, json, os, logging, uuid
from typing import Dict, List, Tuple

# --- Genel ---
ssl._create_default_https_context = ssl._create_unverified_context
# (en başlara)
try:
    if not st.session_state.get("_root_page_config_done", False):
        st.set_page_config(page_title="Çok Dilli Çeviri", layout="wide")
        st.session_state["_root_page_config_done"] = True
except Exception:
    pass

# LocaleId ↔ dil kodu
LOCALE_MAP: Dict[int, str] = {
    1: "en",   # İngilizce
    20: "az",  # Azerbaycanca
    22: "pt",  # Portekizce
    23: "es",  # İspanyolca
    2: "tr"    # Türkçe
}
LANGUAGE_NAMES = {"en": "İngilizce", "az": "Azerbaycanca", "pt": "Portekizce", "es": "İspanyolca", "tr": "Türkçe"}
ALL_LANGUAGES = ["Tümü"] + [name for code, name in LANGUAGE_NAMES.items() if code != "tr"]
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SPECIAL_TRANSLATIONS_FILE = os.path.join(BASE_DIR, "special_translations.json")

# ---------- Yardımcılar ----------
def _safe_int(x) -> int:
    try:
        # Excel bazen 2.0/2.00 döndürebilir; string de olabilir
        return int(float(str(x).strip()))
    except Exception:
        return -1

@st.cache_resource(show_spinner=False)
def _get_translators() -> Dict[str, GoogleTranslator]:
    # tr hariç hedefler
    return {code: GoogleTranslator(source="auto", target=code) for code in LOCALE_MAP.values() if code != "tr"}

TRANSLATORS = _get_translators()

def try_fallback(seg: str, lang: str) -> str:
    try:
        return MyMemoryTranslator(source="auto", target=lang).translate(seg)
    except Exception as e:
        logging.warning(f"Fallback hatası: {e}")
        return seg

def translate_segment(seg: str, lang: str) -> str:
    if lang == "tr" or not seg:
        return seg
    try:
        return TRANSLATORS[lang].translate(seg)
    except exceptions.TranslationNotFound:
        return try_fallback(seg, lang)
    except Exception:
        return seg

def load_special_translations() -> Dict[str, Dict[str, str]]:
    if not os.path.exists(SPECIAL_TRANSLATIONS_FILE):
        return {}
    with open(SPECIAL_TRANSLATIONS_FILE, "r", encoding="utf-8") as f:
        raw = json.load(f)
    out: Dict[str, Dict[str, str]] = {}
    for visible_name, mappings in raw.items():
        if visible_name == "Tümü":
            out["all"] = mappings
        else:
            code = next((c for c, n in LANGUAGE_NAMES.items() if n == visible_name), None)
            if code:
                out[code] = mappings
    return out

def save_special_translations(dic: Dict[str, Dict[str, str]]):
    save_dict = {}
    for key, val in dic.items():
        if key == "all":
            save_dict["Tümü"] = val
        else:
            save_dict[LANGUAGE_NAMES.get(key, key)] = val
    with open(SPECIAL_TRANSLATIONS_FILE, "w", encoding="utf-8") as f:
        json.dump(save_dict, f, ensure_ascii=False, indent=4)

SPECIAL_TRANSLATIONS = load_special_translations()

def translate_text(text: str, target_lang: str, exclude_list: List[str] | None, special_map: Dict[str, Dict[str, str]] | None) -> str:
    if text is None:
        return ""
    s = str(text)
    if not s.strip():
        return s

    punct = r"[.,!?:;()\[\]{}\"'’“”]"
    segments = re.split(rf"({punct}|\n)", s)
    out = []

    for seg in segments:
        if seg == "\n" or re.match(punct, seg):
            out.append(seg); continue

        placeholder_map: Dict[str, str] = {}

        # özel terimler
        if special_map:
            combined = {}
            combined.update(special_map.get("all", {}))
            combined.update(special_map.get(target_lang, {}))
            for idx, (orig, mapped) in enumerate(combined.items()):
                if not orig:
                    continue
                ph = f"\uF100{idx}\uF100"
                placeholder_map[ph] = mapped
                seg = re.sub(rf"\b{re.escape(orig)}\b", ph, seg)

        # hariç tutulanlar
        if exclude_list:
            for i, w in enumerate(exclude_list):
                if not w:
                    continue
                ph2 = f"\uF200{i}\uF200"
                placeholder_map[ph2] = w
                seg = seg.replace(w, ph2)

        tr = translate_segment(seg, target_lang)

        for ph, original in placeholder_map.items():
            tr = tr.replace(ph, original)

        out.append(tr)

    return "".join(out)

def read_excel_with_styles(file) -> Tuple:
    # Yalnızca xlsx destekliyoruz (openpyxl)
    wb = load_workbook(file, data_only=False)
    ws = wb.active

    first_row = [cell.value for cell in ws[1]]
    headers: List[str] = []
    for idx, val in enumerate(first_row):
        headers.append(str(val).strip().replace("\n", "").replace("\t", "") if val else f"column_{idx+1}")

    req = {"LocaleId", "TranslationKey", "TranslationValue"}
    missing = req - set(headers)
    if missing:
        raise KeyError(f"Eksik başlık(lar): {', '.join(sorted(missing))}")

    data = {h: [] for h in headers}
    for row in ws.iter_rows(min_row=2):
        for h, cell in zip(headers, row):
            data[h].append(cell.value)
    df = pd.DataFrame(data)
    return wb, ws, df, headers

def apply_translation(df: pd.DataFrame, exclude_list: List[str], az_as_tr: bool = False) -> pd.DataFrame:
    # LocaleId -> int normalize
    df = df.copy()
    df["LocaleId"] = df["LocaleId"].apply(_safe_int)
    tr_df = df[df["LocaleId"] == 2].reset_index(drop=True)

    if tr_df.empty:
        raise ValueError("Türkçe (LocaleId=2) satırı bulunamadı. Kaynak satır gerekli.")

    rows = []
    total = len(tr_df)
    prog = st.progress(0.0, text="Çeviri hazırlanıyor...")

    for i, r in tr_df.iterrows():
        key = r["TranslationKey"]
        base = r["TranslationValue"]
        rows.append({"LocaleId": 2, "TranslationKey": key, "TranslationValue": base})

        for lid, code in LOCALE_MAP.items():
            if lid == 2:
                continue
            if az_as_tr and lid == 20:
                tr_text = base
            else:
                tr_text = translate_text(str(base or ""), code, exclude_list, SPECIAL_TRANSLATIONS)
            rows.append({"LocaleId": lid, "TranslationKey": key, "TranslationValue": tr_text})

        prog.progress((i + 1) / max(1, total), text=f"Çeviri: {i+1}/{total}")

    res = pd.DataFrame(rows).sort_values(["TranslationKey", "LocaleId"], ascending=[True, True]).reset_index(drop=True)
    return res

# ---------- UI ----------
def main():
    st.title("🌐 Panel Çok Dilli Çeviri")
    st.caption("Kaynak: Türkçe (LocaleId=2). Hedef diller LOCALE_MAP’e göre oluşturulur.")

    # --- Özel çeviri yönetimi ---
    with st.expander("🧩 Özel Çeviri Yönetimi", expanded=False):
        target_name = st.selectbox("Hedef Dil", ALL_LANGUAGES, index=0)
        key_in = st.text_input("Türkçesi (kelime/cümle)")
        val_in = st.text_input("Çeviri")
        if st.button("➕ Ekle"):
            if key_in and val_in:
                code_key = "all" if target_name == "Tümü" else next((c for c, n in LANGUAGE_NAMES.items() if n == target_name), None)
                if code_key:
                    SPECIAL_TRANSLATIONS.setdefault(code_key, {})[key_in] = val_in
                    save_special_translations(SPECIAL_TRANSLATIONS)
                    st.success("Özel çeviri kaydedildi.")

        # Listele + sil
        any_shown = False
        for code, mappings in list(SPECIAL_TRANSLATIONS.items()):
            if mappings:
                any_shown = True
                lang_name = "Tümü" if code == "all" else LANGUAGE_NAMES.get(code, code)
                st.markdown(f"**{lang_name}**")
                for orig, trans in list(mappings.items()):
                    c1, c2, c3 = st.columns([3, 3, 1])
                    c1.write(orig)
                    c2.write(trans)
                    if c3.button("🗑️", key=f"del_{code}_{orig}"):
                        SPECIAL_TRANSLATIONS[code].pop(orig, None)
                        if not SPECIAL_TRANSLATIONS[code]:
                            SPECIAL_TRANSLATIONS.pop(code, None)
                        save_special_translations(SPECIAL_TRANSLATIONS)
                        st.success(f"Kaldırıldı: [{lang_name}] {orig}")
        if not any_shown:
            st.info("Özel çeviri ekli değil.")

    # --- Dosya ve ayarlar ---
    uploaded_file = st.file_uploader("Excel yükle (xlsx)", type=["xlsx"])
    az_as_tr = st.checkbox("Azerbaycancayı Türkçe olarak yaz", value=False)
    exclude_list: List[str] = []
    if uploaded_file:
        exclude_input = st.text_area("Çevrilmeyecek kelimeler (virgül veya satır sonu ile ayırın)")
        exclude_list = [w.strip() for w in re.split(r"[,\n]", exclude_input or "") if w.strip()]

    # --- Çalıştır ---
    if st.button("🚀 Çevir ve Önizle"):
        if not uploaded_file:
            st.warning("Önce bir Excel dosyası yükleyin (.xlsx).")
            st.stop()
        try:
            wb, ws, df, headers = read_excel_with_styles(uploaded_file)
        except Exception as e:
            st.error(f"Excel okunamadı: {e}")
            st.stop()

        try:
            result = apply_translation(df, exclude_list, az_as_tr)
        except Exception as e:
            st.error(f"Çeviri hatası: {e}")
            st.stop()

        st.subheader("Önizleme")
        st.dataframe(result, use_container_width=True)

        # Çıktı sayfasını oluştur ve yaz
        out_name = "Çevrilmiş Veri"
        if out_name in [ws.title for ws in wb.worksheets]:
            del wb[out_name]
        out_ws = wb.create_sheet(out_name)

        cols = ["TranslationKey", "LocaleId", "TranslationValue"]
        for c_idx, col in enumerate(cols, start=1):
            out_ws.cell(row=1, column=c_idx, value=col)
        for r_idx, row in enumerate(result.itertuples(index=False), start=2):
            out_ws.cell(row=r_idx, column=1, value=row.TranslationKey)
            out_ws.cell(row=r_idx, column=2, value=row.LocaleId)
            out_ws.cell(row=r_idx, column=3, value=row.TranslationValue)

        # İndirme
        output_name = f"translated_{uuid.uuid4().hex}.xlsx"
        wb.save(output_name)
        with open(output_name, "rb") as f:
            st.download_button("⬇️ Excel indir", f, file_name=output_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# --- birleşik panel uyumlu wrapper ---
def run(embedded: bool = False):
    # main() zaten st.set_page_config çağırıyor; ana panelde sorun çıkmasın diye try/except
    if not embedded:
        try:
            st.set_page_config(page_title="Çok Dilli Çeviri", layout="wide")
        except Exception:
            pass
    main()

if __name__ == "__main__":
    run(False)
